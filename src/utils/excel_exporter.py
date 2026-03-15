<<<<<<< HEAD
"""
Excel 导出器模块
负责将采集到的商品数据导出为结构化的 Excel 表格，
支持格式美化、列宽自适应和中文表头。
"""

import os
import pandas as pd
from datetime import datetime


class ExcelExporter:
    """
    商品数据 Excel 导出器。
    将商品详情列表导出为格式化的 .xlsx 文件，
    包含冻结首行、自动列宽、颜色样式等特性。
    """

    # 中文表头映射（数据字段 → 显示名称）
    COLUMN_MAP = {
        "序号": "序号",
        "商品名称": "商品名称",
        "价格(元)": "价格(元)",
        "起批量": "起批量",
        "供应商": "供应商",
        "发货地": "发货地",
        "材质": "材质",
        "颜色/规格": "颜色/规格",
        "来源链接": "来源链接",
        "图片数量": "图片数量",
        "采集时间": "采集时间",
    }

    def __init__(self, output_dir="1688_products"):
        """
        初始化导出器。
        
        参数:
            output_dir: 输出目录路径
        """
        self.output_dir = output_dir
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

    def export(self, products, filename=None):
        """
        将商品列表导出为 Excel 文件。
        
        参数:
            products: 商品信息字典列表
            filename: 输出文件名（不含路径），默认自动生成带时间戳的文件名
            
        返回:
            导出的 Excel 文件完整路径
        """
        if not products:
            print("[ExcelExporter] 警告：商品列表为空，跳过导出。")
            return None

        # 构造 DataFrame 数据
        rows = []
        for idx, product in enumerate(products, 1):
            attributes = product.get("attributes", {})
            row = {
                "序号": idx,
                "商品名称": product.get("title", "未知商品"),
                "价格(元)": product.get("price", "N/A"),
                "起批量": attributes.get("起批量", attributes.get("最小起订量", "N/A")),
                "供应商": attributes.get("供应商", attributes.get("公司名称", "N/A")),
                "发货地": attributes.get("发货地", attributes.get("所在地区", "N/A")),
                "材质": attributes.get("材质", attributes.get("面料", "N/A")),
                "颜色/规格": attributes.get("颜色", attributes.get("规格", "N/A")),
                "来源链接": product.get("url", "N/A"),
                "图片数量": len(product.get("images", [])),
                "采集时间": datetime.now().strftime("%Y-%m-%d %H:%M"),
            }
            rows.append(row)

        df = pd.DataFrame(rows)

        # 生成文件名
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"1688_商品采集报表_{timestamp}.xlsx"

        filepath = os.path.join(self.output_dir, filename)

        # 使用 openpyxl 引擎写入并格式化
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="商品详情", index=False)
            
            # 获取工作表对象用于格式美化
            workbook = writer.book
            worksheet = writer.sheets["商品详情"]
            
            # ---------- 格式化样式 ----------
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            # 表头样式：深蓝背景 + 白色加粗字体
            header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # 数据行样式
            data_font = Font(name="微软雅黑", size=10)
            data_alignment = Alignment(vertical="center", wrap_text=True)
            
            # 交替行颜色（斑马纹）
            even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
            
            # 边框样式
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # 应用表头样式
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # 应用数据行样式
            for row_idx in range(2, len(df) + 2):
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
                    # 斑马纹：偶数行着色
                    if row_idx % 2 == 0:
                        cell.fill = even_fill

            # ---------- 自适应列宽 ----------
            for col_idx, col_name in enumerate(df.columns, 1):
                # 计算最大内容宽度（中文字符按 2 倍宽计算）
                max_length = len(str(col_name))
                for row_idx in range(2, len(df) + 2):
                    cell_value = str(worksheet.cell(row=row_idx, column=col_idx).value or "")
                    # 中文字符宽度估算
                    char_count = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                    max_length = max(max_length, char_count)
                
                # 设置列宽，添加 padding，并限制最大宽度
                adjusted_width = min(max_length + 4, 50)
                worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = adjusted_width

            # 冻结首行（滚动时表头始终可见）
            worksheet.freeze_panes = "A2"
            
            # 设置行高
            worksheet.row_dimensions[1].height = 28  # 表头行高
            for row_idx in range(2, len(df) + 2):
                worksheet.row_dimensions[row_idx].height = 22

        print(f"[ExcelExporter] ✅ Excel 报表已成功导出: {filepath}")
        return filepath
=======
# -*- coding: utf-8 -*-
import pandas as pd
import os

class ExcelExporter:
    """Excel导出工具"""
    
    def export(self, data, filename):
        """
        导出数据到Excel文件
        
        Args:
            data (list): 包含字典的列表数据
            filename (str): 输出文件名或路径
            
        Returns:
            str: 导出的文件绝对路径
        """
        if not data:
            print("没有数据可导出")
            return None
            
        try:
            # 创建DataFrame
            df = pd.DataFrame(data)
            
            # 处理输出路径
            output_path = os.path.abspath(filename)
            output_dir = os.path.dirname(output_path)
            
            # 确保目录存在
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # 导出Excel
            df.to_excel(output_path, index=False)
            
            return output_path
            
        except Exception as e:
            print(f"导出Excel失败: {e}")
            raise e
>>>>>>> release/v1.3.3-hotfix
