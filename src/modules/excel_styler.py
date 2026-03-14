"""
Excel样式模块 - Excel格式化样式定义
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import Dict


class ExcelStyler:
    """
    Excel 样式器

    提供预定义的 Excel 样式
    """

    # 表头样式
    HEADER_STYLE = {
        "font": Font(name="微软雅黑", bold=True, color="FFFFFF", size=11),
        "fill": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }

    # 数据行样式
    DATA_STYLE = {
        "font": Font(name="微软雅黑", size=10),
        "alignment": Alignment(vertical="center", wrap_text=True),
    }

    # 交替行颜色
    EVEN_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # 边框样式
    BORDER_STYLE = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def __init__(self):
        pass

    def get_header_style(self) -> Dict:
        """获取表头样式"""
        return dict(self.HEADER_STYLE)

    def get_data_style(self) -> Dict:
        """获取数据行样式"""
        return dict(self.DATA_STYLE)

    def get_even_row_style(self) -> Dict:
        """获取交替行样式"""
        style = dict(self.DATA_STYLE)
        style["fill"] = self.EVEN_ROW_FILL
        return style

    def apply_header_style(self, cell):
        """应用表头样式到单元格"""
        cell.font = self.HEADER_STYLE["font"]
        cell.fill = self.HEADER_STYLE["fill"]
        cell.alignment = self.HEADER_STYLE["alignment"]
        cell.border = self.BORDER_STYLE

    def apply_data_style(self, cell, is_even_row: bool = False):
        """应用数据样式到单元格"""
        cell.font = self.DATA_STYLE["font"]
        cell.alignment = self.DATA_STYLE["alignment"]
        cell.border = self.BORDER_STYLE

        if is_even_row:
            cell.fill = self.EVEN_ROW_FILL

    @staticmethod
    def calculate_column_width(content: str, padding: int = 4) -> int:
        """
        计算列宽

        Args:
            content: 内容
            padding: 边距

        Returns:
            建议的列宽
        """
        # 中文字符按2倍宽计算
        char_count = sum(2 if ord(c) > 127 else 1 for c in str(content))
        return min(char_count + padding, 50)


# 预定义颜色
class Colors:
    """预定义颜色"""
    HEADER_BLUE = "1F4E79"
    ODD_ROW = "FFFFFF"
    EVEN_ROW = "D6E4F0"
    SUCCESS = "C6EFCE"
    WARNING = "FFEB9C"
    ERROR = "FFC7CE"


# 便捷函数
def get_styler() -> ExcelStyler:
    """获取样式器实例"""
    return ExcelStyler()


if __name__ == "__main__":
    # 测试
    styler = ExcelStyler()

    print("样式测试:")
    print(f"表头字体: {styler.HEADER_STYLE['font'].name}")
    print(f"表头颜色: {styler.HEADER_STYLE['fill'].start_color}")
    print(f"数据字体: {styler.DATA_STYLE['font'].name}")

    # 列宽计算测试
    test_contents = [
        "中文测试内容",
        "English content",
        "混合内容Test123",
    ]

    for content in test_contents:
        width = styler.calculate_column_width(content)
        print(f"'{content}' -> 宽度 {width}")
